# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import unicodedata
import re
import altair as alt
import plotly.express as px
import json
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.enums import TA_LEFT, TA_CENTER

st.set_page_config(
    page_title="Ficha de Cooperaci\u00f3n Internacional | APC Colombia",
    layout="wide",
    initial_sidebar_state="collapsed"
)
st.markdown("""
    <style>
    :root { color-scheme: light; }
    </style>
""", unsafe_allow_html=True)

FILE = "Ficha_territorial.xlsm"
FILE_SECTORES = "Ficha sectores.xlsx"
GEO_FILE = "Colombia_geo.json"
LOGO_APC = "logo_apc.png"
LOGO_SNCIC = "logo_sncic.png"

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Montserrat:wght@400;500;600;700;800&family=Source+Sans+3:wght@400;600&display=swap');

:root {
    --apc-blue:      #003087;
    --apc-blue-mid:  #004BB4;
    --apc-red:       #C8102E;
    --apc-yellow:    #F5A623;
    --apc-light:     #EEF3FB;
    --apc-gray:      #F7F8FA;
    --apc-gray-mid:  #EAECF0;
    --apc-border:    #D1D9E6;
    --apc-text:      #1C2B4A;
    --apc-muted:     #5A6A85;
    --apc-white:     #FFFFFF;
}

html, body, [class*="css"] {
    font-family: 'Source Sans 3', sans-serif;
    color: var(--apc-text);
    background-color: #F7F8FA;
}

/* \u2500\u2500 Header \u2500\u2500 */
.apc-header {
    background: #003087;
    padding: 1.2rem 2.2rem 1rem 2.2rem;
    margin-bottom: 0;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 4px solid var(--apc-red);
}
.apc-header-title {
    color: white;
    font-family: 'Montserrat', sans-serif;
    font-size: 1.5rem;
    font-weight: 700;
    letter-spacing: 0.2px;
    margin: 0;
}
.apc-header-subtitle {
    color: rgba(255,255,255,0.75);
    font-size: 0.82rem;
    font-weight: 400;
    margin-top: 3px;
    font-family: 'Source Sans 3', sans-serif;
    letter-spacing: 0.3px;
}
.apc-flag-bar {
    height: 5px;
    background: linear-gradient(90deg, #F5A623 33.3%, #003087 33.3% 66.6%, #C8102E 66.6%);
    margin-bottom: 1.5rem;
}

/* \u2500\u2500 Selector card \u2500\u2500 */
.dept-selector-card {
    background: var(--apc-white);
    border-left: 5px solid var(--apc-blue);
    border-radius: 0 8px 8px 0;
    padding: 0.9rem 1.4rem;
    margin-bottom: 1.4rem;
    box-shadow: 0 1px 4px rgba(0,48,135,0.07);
}

/* \u2500\u2500 Banner departamento \u2500\u2500 */
.dept-title-banner {
    background: var(--apc-blue);
    color: white;
    font-family: 'Montserrat', sans-serif;
    font-size: 1.2rem;
    font-weight: 700;
    padding: 0.75rem 1.5rem;
    border-radius: 6px;
    margin-bottom: 1.2rem;
    border-left: 6px solid var(--apc-red);
    letter-spacing: 0.2px;
}

/* \u2500\u2500 Section headers \u2500\u2500 */
.section-header {
    font-family: 'Montserrat', sans-serif;
    font-weight: 700;
    font-size: 0.95rem;
    color: var(--apc-blue);
    text-transform: uppercase;
    letter-spacing: 1px;
    border-bottom: 3px solid var(--apc-yellow);
    padding-bottom: 6px;
    margin: 2rem 0 1rem 0;
}

/* \u2500\u2500 M\u00e9tricas \u2500\u2500 */
div[data-testid="stMetric"] {
    background: var(--apc-white);
    border: 1px solid var(--apc-border);
    border-left: 5px solid var(--apc-blue);
    border-radius: 6px;
    padding: 1rem 1.1rem !important;
    box-shadow: 0 1px 6px rgba(0,48,135,0.06);
    transition: box-shadow 0.2s, border-left-color 0.2s;
}
div[data-testid="stMetric"]:hover {
    box-shadow: 0 4px 14px rgba(0,48,135,0.12);
    border-left-color: var(--apc-red);
}
div[data-testid="stMetricLabel"] {
    font-family: 'Montserrat', sans-serif;
    font-weight: 600;
    font-size: 0.72rem;
    color: var(--apc-muted) !important;
    text-transform: uppercase;
    letter-spacing: 0.7px;
}
div[data-testid="stMetricValue"] {
    font-family: 'Montserrat', sans-serif;
    font-weight: 700;
    font-size: 1.6rem !important;
    color: var(--apc-blue) !important;
}

/* \u2500\u2500 Tabs \u2500\u2500 */
button[data-baseweb="tab"] {
    font-family: 'Montserrat', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.85rem !important;
    letter-spacing: 0.4px;
    text-transform: uppercase;
    color: var(--apc-muted) !important;
}
button[data-baseweb="tab"][aria-selected="true"] {
    color: var(--apc-blue) !important;
}
div[data-baseweb="tab-highlight"] {
    background-color: var(--apc-red) !important;
    height: 3px !important;
}
div[data-baseweb="tab-border"] {
    background-color: var(--apc-border) !important;
}

/* \u2500\u2500 Dataframes \u2500\u2500 */
div[data-testid="stDataFrame"] {
    border-radius: 6px;
    overflow: hidden;
    border: 1px solid var(--apc-border);
    box-shadow: 0 1px 4px rgba(0,0,0,0.04);
}

/* \u2500\u2500 Botones descarga \u2500\u2500 */
div[data-testid="stDownloadButton"] button {
    background: var(--apc-blue) !important;
    color: white !important;
    border: none !important;
    border-radius: 4px !important;
    font-family: 'Montserrat', sans-serif !important;
    font-weight: 600 !important;
    font-size: 0.82rem !important;
    letter-spacing: 0.5px;
    text-transform: uppercase;
    padding: 0.5rem 1.4rem !important;
    transition: background 0.2s !important;
}
div[data-testid="stDownloadButton"] button:hover {
    background: var(--apc-red) !important;
}

/* \u2500\u2500 Guia card \u2500\u2500 */
.guia-card {
    background: var(--apc-white);
    border: 1px solid var(--apc-border);
    border-radius: 8px;
    padding: 2rem 2.4rem;
    margin-bottom: 1rem;
    line-height: 1.8;
    font-size: 0.97rem;
    box-shadow: 0 1px 6px rgba(0,48,135,0.05);
}
.guia-card p {
    color: var(--apc-text);
    margin-bottom: 1.1rem;
}
.guia-intro {
    font-family: 'Montserrat', sans-serif;
    font-size: 0.95rem;
    font-weight: 700;
    color: var(--apc-blue);
    background: var(--apc-light);
    border-left: 5px solid var(--apc-red);
    border-radius: 0 6px 6px 0;
    padding: 0.8rem 1.2rem;
    margin-bottom: 1.4rem;
    text-transform: uppercase;
    letter-spacing: 0.5px;
}

/* \u2500\u2500 Footer \u2500\u2500 */
.apc-footer {
    text-align: center;
    color: var(--apc-muted);
    font-size: 0.76rem;
    margin-top: 3rem;
    padding-top: 1rem;
    border-top: 1px solid var(--apc-border);
    letter-spacing: 0.3px;
}

div[data-testid="stMetricLabel"] p {
    font-size: 0.65rem !important;
    line-height: 1.3 !important;
    white-space: normal !important;
}

/* Metrica personalizada con texto largo */
.metric-custom {
    background: var(--apc-white);
    border: 1px solid var(--apc-border);
    border-left: 5px solid var(--apc-blue);
    border-radius: 6px;
    padding: 1rem 1.1rem;
    box-shadow: 0 1px 6px rgba(0,48,135,0.06);
    transition: box-shadow 0.2s, border-left-color 0.2s;
    height: 100%;
}
.metric-custom:hover {
    box-shadow: 0 4px 14px rgba(0,48,135,0.12);
    border-left-color: #C8102E;
}
.metric-custom-label {
    font-family: 'Montserrat', sans-serif;
    font-weight: 600;
    font-size: 0.68rem;
    color: #5A6A85;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    line-height: 1.35;
    margin-bottom: 0.4rem;
}
.metric-custom-value {
    font-family: 'Montserrat', sans-serif;
    font-weight: 700;
    font-size: 1.6rem;
    color: #003087;
    line-height: 1.2;
}
.metric-custom-delta {
    font-family: 'Montserrat', sans-serif;
    font-size: 0.72rem;
    font-weight: 600;
    margin-top: 4px;
}
.metric-custom-sub {
    font-family: 'Source Sans 3', sans-serif;
    font-size: 0.72rem;
    color: #5A6A85;
    margin-top: 2px;
    line-height: 1.4;
}

/* Ocultar barra Streamlit */
div[data-testid="stToolbar"],
div[data-testid="stDecoration"],
header[data-testid="stHeader"] {
    display: none !important;
}

/* \u2500\u2500 Logo header \u2500\u2500 */
.app-header {
    background: #003087;
    padding: 0.7rem 2rem;
    display: flex;
    align-items: center;
    justify-content: space-between;
    border-bottom: 4px solid #C8102E;
    margin-bottom: 0;
}
.app-header-center {
    text-align: center;
    flex: 1;
    padding: 0 1rem;
}
.app-header-title {
    color: white;
    font-family: 'Montserrat', sans-serif;
    font-size: 1rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: 0.3px;
}
.app-header-sub {
    color: rgba(255,255,255,0.72);
    font-size: 0.75rem;
    margin-top: 2px;
    font-family: 'Source Sans 3', sans-serif;
}
.logo-side {
    width: 120px;
    display: flex;
    align-items: center;
    justify-content: center;
}

/* \u2500\u2500 Panel general cards \u2500\u2500 */
.panel-stat {
    background: white;
    border: 1px solid #D1D9E6;
    border-left: 5px solid #003087;
    border-radius: 6px;
    padding: 1rem 1.2rem;
    margin-bottom: 0.5rem;
    box-shadow: 0 1px 4px rgba(0,48,135,0.06);
}
.panel-stat-label {
    font-family: 'Montserrat', sans-serif;
    font-size: 0.68rem;
    font-weight: 600;
    color: #5A6A85;
    text-transform: uppercase;
    letter-spacing: 0.6px;
    margin-bottom: 4px;
}
.panel-stat-value {
    font-family: 'Montserrat', sans-serif;
    font-size: 1.6rem;
    font-weight: 700;
    color: #003087;
    line-height: 1.1;
}
.panel-stat-sub {
    font-size: 0.72rem;
    color: #5A6A85;
    margin-top: 3px;
}
.panel-stat-red { border-left-color: #C8102E; }
.panel-stat-yellow { border-left-color: #F5A623; }

/* \u2500\u2500 Section divider with icon \u2500\u2500 */
.section-title {
    font-family: 'Montserrat', sans-serif;
    font-weight: 700;
    font-size: 0.9rem;
    color: #003087;
    text-transform: uppercase;
    letter-spacing: 1px;
    border-bottom: 3px solid #F5A623;
    padding-bottom: 6px;
    margin: 1.8rem 0 1rem 0;
}

/* \u2500\u2500 Selector pill \u2500\u2500 */
div[data-testid="stSelectbox"] > div > div {
    border-radius: 6px !important;
    border-color: #D1D9E6 !important;
}

/* \u2500\u2500 Map container \u2500\u2500 */
.map-title {
    font-family: 'Montserrat', sans-serif;
    font-weight: 600;
    font-size: 0.8rem;
    color: #5A6A85;
    text-transform: uppercase;
    letter-spacing: 0.5px;
    margin-bottom: 6px;
}

/* Fondo blanco forzado */
.stApp {
    background-color: #FFFFFF !important;
}
[data-testid="stAppViewContainer"] {
    background-color: #FFFFFF !important;
}
[data-testid="stMain"] {
    background-color: #FFFFFF !important;
}

/* Forzar modo claro en Mac */
:root {
    color-scheme: light only;
}
body {
    background-color: #FFFFFF !important;
}

/* Textos visibles */
[data-testid="stMarkdownContainer"] p,
[data-testid="stMarkdownContainer"] li,
[data-testid="stCaptionContainer"] p,
[data-testid="stRadio"] label span,
[data-testid="stMetricLabel"] p {
    color: #1C2B4A !important;
}

/* Fondo blanco en gr\u00e1ficas - Mac */
[data-testid="stArrowVegaLiteChart"] > div,
[data-testid="stDataFrame"] > div,
.js-plotly-plot,
.plotly {
    background-color: #FFFFFF !important;
}

/* Texto botones */
div[data-testid="stDownloadButton"] button,
div[data-testid="stDownloadButton"] button p,
div[data-testid="stDownloadButton"] button span {
    color: white !important;
}


</style>
""", unsafe_allow_html=True)

def norm_text(x):
    if x is None:
        return ""
    s = str(x).strip().upper()
    s = "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")
    s = re.sub(r"[^A-Z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def format_usd(n):
    try:
        n = float(n)
    except Exception:
        return ""
    return "USD " + f"{n:,.0f}".replace(",", ".")


def format_cop(n):
    try:
        n = float(n)
    except Exception:
        return ""
    return "$ " + f"{n:,.0f}".replace(",", ".")


def get_col(row, *names):
    """Busca una columna por varios nombres posibles (con y sin tilde)."""
    for name in names:
        val = row.get(name, None)
        if val is not None and str(val).strip() not in ("", "None", "nan"):
            return val
    return ""



@st.cache_data
def read_named_table(file_path: str, table_name: str) -> pd.DataFrame:
    wb = load_workbook(file_path, data_only=True, keep_vba=True)
    for ws in wb.worksheets:
        if table_name in ws.tables:
            tbl = ws.tables[table_name]
            min_col, min_row, max_col, max_row = range_boundaries(tbl.ref)
            data = []
            for row in ws.iter_rows(
                min_row=min_row, max_row=max_row,
                min_col=min_col, max_col=max_col,
                values_only=True
            ):
                data.append(list(row))
            header = data[0]
            rows = data[1:]
            return pd.DataFrame(rows, columns=header)
    raise KeyError(f"No encontre la tabla: {table_name}")


@st.cache_data
def load_data():
    infogeneral = read_named_table(FILE, "infogeneral")
    plan = read_named_table(FILE, "plan")
    ciclope = read_named_table(FILE, "Tabla7")  # nuevo corte 2026-2
    ciclope_ant = read_named_table(FILE, "ciclope20261")  # comparativo 2026-1
    wb_css = load_workbook(FILE, data_only=True, keep_vba=True)
    ws_css = wb_css["CSS"]
    css_data = [list(row) for row in ws_css.iter_rows(values_only=True)]
    css = pd.DataFrame(css_data[1:], columns=css_data[0])
    for c in css.columns:
        if css[c].dtype == "object":
            css[c] = css[c].astype(str).str.strip()
    colcol = read_named_table(FILE, "colcol")
    contrapartidas = read_named_table(FILE, "contrapartidas")
    contrapartidas.columns = [str(c).strip().strip("'") for c in contrapartidas.columns]
    proyectos = read_named_table(FILE, "Tabla7")

    for df in [infogeneral, plan, ciclope, ciclope_ant, colcol, contrapartidas, proyectos]:
        for c in df.columns:
            if df[c].dtype == "object":
                df[c] = df[c].astype(str).str.strip()

    if "VALOR APORTE (USD)" in ciclope.columns:
        ciclope["VALOR APORTE (USD)"] = pd.to_numeric(
            ciclope["VALOR APORTE (USD)"], errors="coerce"
        ).fillna(0)
    if "VALOR APORTE (USD)" in ciclope_ant.columns:
        ciclope_ant["VALOR APORTE (USD)"] = pd.to_numeric(
            ciclope_ant["VALOR APORTE (USD)"], errors="coerce"
        ).fillna(0)

    return infogeneral, plan, ciclope, ciclope_ant, colcol, contrapartidas, proyectos, css


@st.cache_data
def load_sectores():
    """Carga los datos de la ficha sectorial."""
    info_s = pd.read_excel(FILE_SECTORES, sheet_name="INFO GENERAL")
    aod_s = pd.read_excel(FILE_SECTORES, sheet_name="AOD202602")
    aod_s_ant = pd.read_excel(FILE_SECTORES, sheet_name="AOD202601")
    css_s = pd.read_excel(FILE_SECTORES, sheet_name="CSS22026")
    colcol_s = pd.read_excel(FILE_SECTORES, sheet_name="COLCOL")

    for df in [info_s, aod_s, css_s, colcol_s]:
        for c in df.columns:
            if df[c].dtype == "object":
                df[c] = df[c].astype(str).str.strip()

    if "VALOR APORTE (USD)" in aod_s.columns:
        aod_s["VALOR APORTE (USD)"] = pd.to_numeric(
            aod_s["VALOR APORTE (USD)"], errors="coerce"
        ).fillna(0)
    if "VALOR APORTE (USD)" in aod_s_ant.columns:
        aod_s_ant["VALOR APORTE (USD)"] = pd.to_numeric(
            aod_s_ant["VALOR APORTE (USD)"], errors="coerce"
        ).fillna(0)

    return info_s, aod_s, aod_s_ant, css_s, colcol_s


@st.cache_data
def load_geo():
    with open(GEO_FILE, encoding="utf-8") as f:
        return json.load(f)


def make_map(geo, dept_values, selected_dept=None):
    """Genera mapa coropl\u00e9tico interactivo de Colombia."""
    names = [f["properties"]["NOMBRE_DPT"] for f in geo["features"]]
    df_map = pd.DataFrame({"dept_geo": names})
    df_map["value"] = df_map["dept_geo"].map(
        lambda x: dept_values.get(norm_text(x), 0)
    )
    df_map["selected"] = df_map["dept_geo"].map(
        lambda x: 1 if selected_dept and norm_text(x) == norm_text(selected_dept) else 0
    )

    fig = px.choropleth(
        df_map,
        geojson=geo,
        locations="dept_geo",
        featureidkey="properties.NOMBRE_DPT",
        color="value",
        color_continuous_scale=[
            [0, "#EEF3FB"],
            [0.2, "#A8C4E8"],
            [0.5, "#4A90D9"],
            [0.8, "#1565C0"],
            [1.0, "#003087"]
        ],
        hover_name="dept_geo",
        hover_data={"value": True, "selected": False, "dept_geo": False},
        labels={"value": "Intervenciones"},
    )
    fig.update_geos(
        fitbounds="locations",
        visible=False,
        bgcolor="rgba(0,0,0,0)"
    )
    fig.update_layout(
        margin={"r": 0, "t": 0, "l": 0, "b": 0},
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        height=480,
        coloraxis=dict(
            colorbar=dict(
                title=dict(text="Intervenciones", font=dict(size=9)),
                thickness=12,
                len=0.6,
                tickfont=dict(size=9)
            )
        )
    )

    # Highlight selected dept
    if selected_dept:
        selected_features = [
            f for f in geo["features"]
            if norm_text(f["properties"]["NOMBRE_DPT"]) == norm_text(selected_dept)
        ]
        if selected_features:
            import plotly.graph_objects as go
            sel_geo = {"type": "FeatureCollection", "features": selected_features}
            fig.add_trace(go.Choropleth(
                geojson=sel_geo,
                locations=[selected_features[0]["properties"]["NOMBRE_DPT"]],
                z=[1],
                featureidkey="properties.NOMBRE_DPT",
                colorscale=[[0, "#C8102E"], [1, "#C8102E"]],
                showscale=False,
                marker_line_color="white",
                marker_line_width=2,
                hoverinfo="skip"
            ))

    return fig


def top_by_sum(df, group_col, value_col, n=5):
    if df.empty or group_col not in df.columns or value_col not in df.columns:
        return pd.DataFrame(columns=[group_col, value_col])
    return (
        df.groupby(group_col, dropna=False)[value_col]
        .sum()
        .sort_values(ascending=False)
        .head(n)
        .reset_index()
    )



ODS_NOMBRES = {
    "ODS 1":  "ODS 1 - Fin de la pobreza",
    "ODS 2":  "ODS 2 - Hambre cero",
    "ODS 3":  "ODS 3 - Salud y bienestar",
    "ODS 4":  "ODS 4 - Educaci\u00f3n de calidad",
    "ODS 5":  "ODS 5 - Igualdad de g\u00e9nero",
    "ODS 6":  "ODS 6 - Agua limpia y saneamiento",
    "ODS 7":  "ODS 7 - Energ\u00eda asequible y no contaminante",
    "ODS 8":  "ODS 8 - Trabajo decente y crecimiento econ\u00f3mico",
    "ODS 9":  "ODS 9 - Industria, innovaci\u00f3n e infraestructura",
    "ODS 10": "ODS 10 - Reducci\u00f3n de las desigualdades",
    "ODS 11": "ODS 11 - Ciudades y comunidades sostenibles",
    "ODS 12": "ODS 12 - Producci\u00f3n y consumo responsables",
    "ODS 13": "ODS 13 - Acci\u00f3n por el clima",
    "ODS 14": "ODS 14 - Vida submarina",
    "ODS 15": "ODS 15 - Vida de ecosistemas terrestres",
    "ODS 16": "ODS 16 - Paz, justicia e instituciones s\u00f3lidas",
    "ODS 17": "ODS 17 - Alianzas para lograr los objetivos",
}



def to_excel_ficha(info_row, cic_dept, colcol_dept, contr_dept, css_dept=None):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not info_row.empty:
            df_info = info_row.T.reset_index()
            df_info.columns = ["Campo", "Valor"]
            df_info = df_info[
                ~df_info["Campo"].astype(str).str.lower().str.strip().isin(
                    ["porcentaje de avance", "dept_norm"]
                )
            ]
            df_info.to_excel(writer, sheet_name="Informacion General", index=False)
        cic_export = cic_dept.drop(columns=["DEPT_NORM"], errors="ignore")
        cic_export.to_excel(writer, sheet_name="AOD - Ciclope", index=False)
        colcol_dept.copy().to_excel(writer, sheet_name="ColCol", index=False)
        contr_dept.to_excel(writer, sheet_name="Contrapartidas", index=False)
        if css_dept is not None and not css_dept.empty:
            COLS_CSS = [
                "C\u00f3digo", "VIA DE COOPERACION", "MODALIDAD", "PAIS SOCIO", "SEGUNDO OFERENTE",
                "REGION", "NOMBRE DE LA INICIATIVA", "TIPO DE INICIATIVA", "FECHA DE APROBACION",
                "OBJETIVO GENERAL/DESCRIPCION DE LA INICIATIVA", "ESTADO",
                "ENTIDAD(ES) NACIONAL(ES)", "ENTIDAD(ES) EXTRANJERA(S)"
            ]
            cols_css = [c for c in COLS_CSS if c in css_dept.columns]
            css_dept[cols_css].to_excel(writer, sheet_name="Coop Sur Sur", index=False)
    output.seek(0)
    return output.getvalue()


def to_excel_proyectos(df_proj):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_proj.to_excel(writer, sheet_name="Proyectos AOD", index=False)
    output.seek(0)
    return output.getvalue()



def to_pdf_ficha(dept, info_row, cic_dept, colcol_dept, contr_dept, css_dept=None):
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    # Colores institucionales
    AZUL = colors.HexColor("#003087")
    AZUL_CLARO = colors.HexColor("#E8F0FE")
    GRIS = colors.HexColor("#F5F7FA")
    GRIS_BORDE = colors.HexColor("#D0D9EA")
    BLANCO = colors.white

    styles = getSampleStyleSheet()

    # Estilos personalizados
    estilo_titulo = ParagraphStyle("titulo",
        fontName="Helvetica-Bold", fontSize=20, textColor=BLANCO,
        spaceAfter=4, leading=24)
    estilo_subtitulo = ParagraphStyle("subtitulo",
        fontName="Helvetica", fontSize=9, textColor=colors.HexColor("#CBD5E1"),
        spaceAfter=0)
    estilo_dept = ParagraphStyle("dept",
        fontName="Helvetica-Bold", fontSize=15, textColor=BLANCO,
        spaceAfter=0, leading=20)
    estilo_seccion = ParagraphStyle("seccion",
        fontName="Helvetica-Bold", fontSize=9, textColor=AZUL,
        spaceBefore=14, spaceAfter=6)
    estilo_normal = ParagraphStyle("normal",
        fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#1A1A2E"),
        spaceAfter=3, leading=13)
    estilo_label = ParagraphStyle("label",
        fontName="Helvetica-Bold", fontSize=8, textColor=colors.HexColor("#6B7280"),
        spaceAfter=1)
    estilo_valor = ParagraphStyle("valor",
        fontName="Helvetica-Bold", fontSize=14, textColor=AZUL,
        spaceAfter=2)
    estilo_caption = ParagraphStyle("caption",
        fontName="Helvetica", fontSize=7, textColor=colors.HexColor("#6B7280"),
        spaceAfter=4, alignment=TA_CENTER)

    story = []

    # --- HEADER ---
    header_data = [[
        Paragraph("Ficha Territorial", estilo_titulo),
        ""
    ],[
        Paragraph("Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia", estilo_subtitulo),
        Paragraph("APC-Colombia", ParagraphStyle("badge",
            fontName="Helvetica-Bold", fontSize=9, textColor=BLANCO,
            alignment=TA_CENTER))
    ]]
    header_table = Table(header_data, colWidths=["75%", "25%"])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), AZUL),
        ("ROUNDEDCORNERS", [8,8,8,8]),
        ("TOPPADDING", (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 14),
        ("LEFTPADDING", (0,0), (0,-1), 16),
        ("RIGHTPADDING", (-1,0), (-1,-1), 16),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 10))

    # --- BANNER DEPARTAMENTO ---
    dept_table = Table([[Paragraph(f"\U0001f4cd  {dept}", estilo_dept)]], colWidths=["100%"])
    dept_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), AZUL),
        ("ROUNDEDCORNERS", [8,8,8,8]),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
    ]))
    story.append(dept_table)
    story.append(Spacer(1, 10))

    # --- INFORMACION GENERAL ---
    story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
    story.append(Paragraph("Informaci\u00f3n General", estilo_seccion))

    if not info_row.empty:
        row = info_row.iloc[0]

        # Metricas principales
        capital = str(row.get("Capital", ""))
        municipios = str(row.get("N\u00famero de Municipios", row.get("Numero de Municipios", row.get("Municipios", ""))))
        pob_raw = row.get("Poblaci\u00f3n", row.get("Poblacion", ""))
        try:
            poblacion = f"{int(float(pob_raw)):,}".replace(",", ".")
        except:
            poblacion = str(pob_raw)

        metrics_data = [
            [Paragraph("CAPITAL", estilo_label),
             Paragraph("N\u00daMERO DE MUNICIPIOS", estilo_label),
             Paragraph("POBLACI\u00d3N", estilo_label)],
            [Paragraph(capital, estilo_valor),
             Paragraph(municipios, estilo_valor),
             Paragraph(poblacion, estilo_valor)],
        ]
        metrics_table = Table(metrics_data, colWidths=["33%", "33%", "34%"])
        metrics_table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), GRIS),
            ("BOX", (0,0), (0,-1), 1, GRIS_BORDE),
            ("BOX", (1,0), (1,-1), 1, GRIS_BORDE),
            ("BOX", (2,0), (2,-1), 1, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 8),
            ("BOTTOMPADDING", (0,0), (-1,-1), 8),
            ("LEFTPADDING", (0,0), (-1,-1), 10),
            ("LINEABOVE", (0,0), (-1,0), 3, AZUL),
            ("ROUNDEDCORNERS", [6,6,6,6]),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 8))

        # Registro completo
        story.append(Paragraph("Informaci\u00f3n detallada del departamento", estilo_seccion))
        df_det = info_row.T.reset_index()
        df_det.columns = ["Campo", "Valor"]
        df_det = df_det[~df_det["Campo"].astype(str).str.lower().str.strip().isin(
            ["porcentaje de avance", "dept_norm"])]
        df_det = df_det[df_det["Valor"].astype(str).str.strip().isin(["", "None", "nan"]) == False]

        tabla_info = []
        for _, fila in df_det.iterrows():
            tabla_info.append([
                Paragraph(str(fila["Campo"]), estilo_label),
                Paragraph(str(fila["Valor"]), estilo_normal)
            ])

        if tabla_info:
            t = Table(tabla_info, colWidths=["35%", "65%"])
            t.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), BLANCO),
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [BLANCO, GRIS]),
                ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
                ("TOPPADDING", (0,0), (-1,-1), 5),
                ("BOTTOMPADDING", (0,0), (-1,-1), 5),
                ("LEFTPADDING", (0,0), (-1,-1), 8),
            ]))
            story.append(t)

    story.append(Spacer(1, 10))

    # --- AOD ---
    story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
    story.append(Paragraph("Ayuda Oficial al Desarrollo (AOD)", estilo_seccion))
    story.append(Paragraph(
        "Fuente: C\u00edclope a corte de 26 de marzo de 2026",
        estilo_caption))

    cic = cic_dept.drop(columns=["DEPT_NORM"], errors="ignore")
    intervenciones = cic["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in cic.columns else 0
    cooperantes = cic["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in cic.columns else 0
    municipios_aod = (
        cic["MUNICIPIO"].map(norm_text)
        .pipe(lambda s: s[~s.isin(["NO REPORTA", "SIN INFORMACION", "NO APLICA", ""])])
        .nunique() if "MUNICIPIO" in cic.columns else 0
    )
    total_usd = cic["VALOR APORTE (USD)"].sum() if "VALOR APORTE (USD)" in cic.columns else 0
    total_fmt = "USD " + f"{total_usd:,.0f}".replace(",", ".")

    aod_metrics = [
        [Paragraph("INTERVENCIONES", estilo_label),
         Paragraph("COOPERANTES", estilo_label),
         Paragraph("MUNICIPIOS / \u00c1REAS", estilo_label),
         Paragraph("TOTAL APORTE (USD)", estilo_label)],
        [Paragraph(str(intervenciones), estilo_valor),
         Paragraph(str(cooperantes), estilo_valor),
         Paragraph(str(municipios_aod), estilo_valor),
         Paragraph(total_fmt, ParagraphStyle("valor_small",
             fontName="Helvetica-Bold", fontSize=10, textColor=AZUL))],
    ]
    t_aod = Table(aod_metrics, colWidths=["25%","25%","25%","25%"])
    t_aod.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), GRIS),
        ("LINEABOVE", (0,0), (-1,0), 3, AZUL),
        ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 10),
    ]))
    story.append(t_aod)
    story.append(Spacer(1, 8))

    # Top cooperantes
    if "NOMBRE ACTOR" in cic.columns and "VALOR APORTE (USD)" in cic.columns:
        top_coop = (cic.groupby("NOMBRE ACTOR")["VALOR APORTE (USD)"]
                    .sum().sort_values(ascending=False).head(5).reset_index())
        story.append(Paragraph("Top 5 cooperantes por aporte estimado (USD)", estilo_seccion))
        coop_data = [[
            Paragraph("COOPERANTE", estilo_label),
            Paragraph("APORTE ESTIMADO", estilo_label)
        ]]
        for _, r in top_coop.iterrows():
            coop_data.append([
                Paragraph(str(r["NOMBRE ACTOR"]), estilo_normal),
                Paragraph("USD " + f"{r['VALOR APORTE (USD)']:,.0f}".replace(",","."), estilo_normal)
            ])
        t_coop = Table(coop_data, colWidths=["70%","30%"])
        t_coop.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ]))
        story.append(t_coop)
        story.append(Spacer(1, 8))

    # Top ODS
    if "ODS" in cic.columns and "VALOR APORTE (USD)" in cic.columns:
        top_ods = (cic.groupby("ODS")["VALOR APORTE (USD)"]
                   .sum().sort_values(ascending=False).head(5).reset_index())
        story.append(Paragraph("Top 5 ODS por aporte estimado (USD)", estilo_seccion))
        ods_data = [[
            Paragraph("ODS", estilo_label),
            Paragraph("APORTE ESTIMADO", estilo_label)
        ]]
        for _, r in top_ods.iterrows():
            nombre_ods = ODS_NOMBRES.get(r["ODS"], r["ODS"])
            ods_data.append([
                Paragraph(nombre_ods, estilo_normal),
                Paragraph("USD " + f"{r['VALOR APORTE (USD)']:,.0f}".replace(",","."), estilo_normal)
            ])
        t_ods = Table(ods_data, colWidths=["70%","30%"])
        t_ods.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ]))
        story.append(t_ods)

    story.append(Spacer(1, 10))

    # --- PROGRAMAS INTERNOS ---
    story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
    story.append(Paragraph("Programas Internos APC-Colombia", estilo_seccion))

    # ColCol
    story.append(Paragraph(f"ColCol - Colombia Ense\u00f1a Colombia ({len(colcol_dept)} registros)", estilo_label))
    if not colcol_dept.empty:
        COLS_COLCOL = [
            "NOMBRE DEL INTERCAMBIO",
            "A\u00d1O DE REALIZACI\u00d3N ",
            "DEPARTAMENTO EN EL QUE SE DESARROLL\u00d3",
            "PRESUPUESTO ESTIMADO APC COLOMBIA",
            "RUBRO ASUMIDO"
        ]
        cols_colcol = [c for c in COLS_COLCOL if c in colcol_dept.columns]
        colcol_show = colcol_dept[cols_colcol].head(30)
        HEADERS_COLCOL = {
            "NOMBRE DEL INTERCAMBIO": "Nombre del intercambio",
            "A\u00d1O DE REALIZACI\u00d3N ": "A\u00f1o",
            "DEPARTAMENTO EN EL QUE SE DESARROLL\u00d3": "Departamento",
            "PRESUPUESTO ESTIMADO APC COLOMBIA": "Presupuesto APC",
            "RUBRO ASUMIDO": "Rubro"
        }
        colcol_show = colcol_show.copy()
        if "PRESUPUESTO ESTIMADO APC COLOMBIA" in colcol_show.columns:
            colcol_show["PRESUPUESTO ESTIMADO APC COLOMBIA"] = (
                pd.to_numeric(colcol_show["PRESUPUESTO ESTIMADO APC COLOMBIA"], errors="coerce")
                .apply(format_cop)
            )
        cc_data = [[Paragraph(HEADERS_COLCOL.get(c, c), estilo_label) for c in colcol_show.columns]]
        for _, r in colcol_show.iterrows():
            cc_data.append([Paragraph(str(v)[:80], estilo_normal) for v in r.values])
        col_widths_cc = ["35%", "8%", "18%", "22%", "17%"][:len(colcol_show.columns)]
        t_cc = Table(cc_data, colWidths=col_widths_cc)
        t_cc.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("FONTSIZE", (0,0), (-1,-1), 8),
        ]))
        story.append(t_cc)
    else:
        story.append(Paragraph("Sin registros para este departamento.", estilo_normal))
    story.append(Spacer(1, 8))

    # Contrapartidas
    story.append(Paragraph(f"Contrapartidas ({len(contr_dept)} registros)", estilo_label))
    if not contr_dept.empty:
        cols_contr = [c for c in contr_dept.columns if c not in ["DEPT_NORM", "Departamento"]]
        contr_show = contr_dept[cols_contr].head(30).copy()
        for col in contr_show.columns:
            if str(col).strip().strip("\'") in ["Monto por APC", "Monto total", "Monto total "]:
                contr_show[col] = pd.to_numeric(contr_show[col], errors="coerce").apply(format_cop)
        ct_data = [[Paragraph(str(c), estilo_label) for c in contr_show.columns]]
        for _, r in contr_show.iterrows():
            ct_data.append([Paragraph(str(v)[:80], estilo_normal) for v in r.values])
        n_cols = len(contr_show.columns)
        col_w = 100 / n_cols
        t_ct = Table(ct_data, colWidths=[f"{col_w}%" for _ in range(n_cols)])
        t_ct.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("WORDWRAP", (0,0), (-1,-1), True),
        ]))
        story.append(t_ct)
    else:
        story.append(Paragraph("Sin registros para este departamento.", estilo_normal))

    # --- COOPERACION SUR-SUR ---
    if css_dept is not None and not css_dept.empty:
        story.append(Spacer(1, 10))
        story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
        story.append(Paragraph("Proyectos de Cooperaci\u00f3n Sur Sur aprobados y vigentes", estilo_seccion))
        story.append(Paragraph(
            f"Datos actualizados a abril de 2026 \u00b7 {len(css_dept)} proyecto(s)",
            estilo_caption))
        COLS_CSS_PDF = [
            "C\u00f3digo", "VIA DE COOPERACION", "PAIS SOCIO", "REGION",
            "NOMBRE DE LA INICIATIVA", "TIPO DE INICIATIVA", "ESTADO",
            "ENTIDAD(ES) NACIONAL(ES)", "ENTIDAD NACIONAL"
        ]
        cols_css_pdf = [c for c in COLS_CSS_PDF if c in css_dept.columns]
        css_show = css_dept[cols_css_pdf].head(30)
        css_data_pdf = [[Paragraph(str(c), estilo_label) for c in css_show.columns]]
        for _, r in css_show.iterrows():
            css_data_pdf.append([Paragraph(str(v)[:80], estilo_normal) for v in r.values])
        n_css = len(css_show.columns)
        col_w_css = 100 / n_css
        t_css = Table(css_data_pdf, colWidths=[f"{col_w_css}%" for _ in range(n_css)])
        t_css.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("FONTSIZE", (0,0), (-1,-1), 7),
        ]))
        story.append(t_css)

    # Footer
    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRIS_BORDE, spaceAfter=4))
    story.append(Paragraph(
        "Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia - APC-Colombia",
        estilo_caption))

    doc.build(story)
    output.seek(0)
    return output.getvalue()


def to_pdf_proyectos(dept, df_proj):
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )

    AZUL = colors.HexColor("#003087")
    AZUL_CLARO = colors.HexColor("#E8F0FE")
    GRIS = colors.HexColor("#F5F7FA")
    GRIS_BORDE = colors.HexColor("#D0D9EA")
    BLANCO = colors.white

    styles = getSampleStyleSheet()
    estilo_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=20,
        textColor=BLANCO, spaceAfter=4, leading=24)
    estilo_subtitulo = ParagraphStyle("subtitulo", fontName="Helvetica", fontSize=9,
        textColor=colors.HexColor("#CBD5E1"), spaceAfter=0)
    estilo_dept = ParagraphStyle("dept", fontName="Helvetica-Bold", fontSize=15,
        textColor=BLANCO, spaceAfter=0, leading=20)
    estilo_label = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=7,
        textColor=colors.HexColor("#6B7280"), spaceAfter=1)
    estilo_normal = ParagraphStyle("normal", fontName="Helvetica", fontSize=8,
        textColor=colors.HexColor("#1A1A2E"), spaceAfter=3, leading=11)
    estilo_caption = ParagraphStyle("caption", fontName="Helvetica", fontSize=7,
        textColor=colors.HexColor("#6B7280"), spaceAfter=4, alignment=TA_CENTER)

    story = []

    # Header
    header_data = [[
        Paragraph("Ficha Territorial", estilo_titulo), ""
    ],[
        Paragraph("Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia", estilo_subtitulo), ""
    ]]
    header_table = Table(header_data, colWidths=["100%", "0%"])
    header_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), AZUL),
        ("TOPPADDING", (0,0), (-1,-1), 14),
        ("BOTTOMPADDING", (0,0), (-1,-1), 14),
        ("LEFTPADDING", (0,0), (-1,-1), 16),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 10))

    # Banner departamento
    dept_table = Table([[Paragraph(f"\U0001f4cd  {dept} \u2014 Proyectos AOD activos", estilo_dept)]], colWidths=["100%"])
    dept_table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), AZUL),
        ("TOPPADDING", (0,0), (-1,-1), 10),
        ("BOTTOMPADDING", (0,0), (-1,-1), 10),
        ("LEFTPADDING", (0,0), (-1,-1), 14),
    ]))
    story.append(dept_table)
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        f"Total proyectos: {len(df_proj)} | Fuente: C\u00edclope a corte de 26 de marzo de 2026",
        estilo_caption))

    # Tabla de proyectos
    COLS = ["NOMBRE INTERVENCION", "FECHA INICIAL", "FECHA FINAL", "MUNICIPIO", "NOMBRE ACTOR"]
    HEADERS = {
        "NOMBRE INTERVENCION": "Nombre de la intervenci\u00f3n",
        "FECHA INICIAL": "Fecha inicial",
        "FECHA FINAL": "Fecha final",
        "MUNICIPIO": "Municipio",
        "NOMBRE ACTOR": "Cooperante"
    }
    COL_WIDTHS = ["32%", "10%", "10%", "16%", "32%"]

    cols_available = [c for c in COLS if c in df_proj.columns]
    widths_available = [COL_WIDTHS[COLS.index(c)] for c in cols_available]

    if not df_proj.empty and cols_available:
        proj_data = [[Paragraph(HEADERS.get(c, c), estilo_label) for c in cols_available]]
        for _, r in df_proj[cols_available].iterrows():
            proj_data.append([Paragraph(str(v)[:120] if v and str(v) != "nan" else "", estilo_normal) for v in r.values])

        t_proj = Table(proj_data, colWidths=widths_available, repeatRows=1)
        t_proj.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING", (0,0), (-1,-1), 6),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
        ]))
        story.append(t_proj)
    else:
        story.append(Paragraph("Sin proyectos para este departamento.", estilo_normal))

    story.append(Spacer(1, 16))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRIS_BORDE, spaceAfter=4))
    story.append(Paragraph(
        "Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia - APC-Colombia",
        estilo_caption))

    doc.build(story)
    output.seek(0)
    return output.getvalue()

def to_pdf_sectorial(sector, info_sector, aod_sector, css_sector, colcol_sector):
    output = BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm
    )
    AZUL = colors.HexColor("#003087")
    AZUL_CLARO = colors.HexColor("#E8F0FE")
    GRIS = colors.HexColor("#F5F7FA")
    GRIS_BORDE = colors.HexColor("#D0D9EA")
    BLANCO = colors.white

    estilo_titulo = ParagraphStyle("titulo", fontName="Helvetica-Bold", fontSize=18,
        textColor=BLANCO, spaceAfter=4, leading=22)
    estilo_subtitulo = ParagraphStyle("subtitulo", fontName="Helvetica", fontSize=8,
        textColor=colors.HexColor("#CBD5E1"), spaceAfter=0)
    estilo_dept = ParagraphStyle("dept", fontName="Helvetica-Bold", fontSize=13,
        textColor=BLANCO, spaceAfter=0, leading=18)
    estilo_seccion = ParagraphStyle("seccion", fontName="Helvetica-Bold", fontSize=9,
        textColor=AZUL, spaceBefore=12, spaceAfter=5)
    estilo_normal = ParagraphStyle("normal", fontName="Helvetica", fontSize=7,
        textColor=colors.HexColor("#1A1A2E"), spaceAfter=2, leading=10)
    estilo_label = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=7,
        textColor=colors.HexColor("#6B7280"), spaceAfter=1)
    estilo_caption = ParagraphStyle("caption", fontName="Helvetica", fontSize=6,
        textColor=colors.HexColor("#6B7280"), spaceAfter=4, alignment=TA_CENTER)

    story = []

    # Header
    hdr = Table([[
        Paragraph("Ficha Sectorial", estilo_titulo),
        Paragraph("APC-Colombia", ParagraphStyle("badge", fontName="Helvetica-Bold",
            fontSize=8, textColor=BLANCO, alignment=TA_CENTER))
    ]], colWidths=["80%", "20%"])
    hdr.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), AZUL),
        ("TOPPADDING", (0,0), (-1,-1), 12),
        ("BOTTOMPADDING", (0,0), (-1,-1), 12),
        ("LEFTPADDING", (0,0), (0,-1), 14),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(hdr)
    story.append(Spacer(1, 8))

    banner = Table([[Paragraph(f"\U0001f3db\ufe0f  {sector}", estilo_dept)]], colWidths=["100%"])
    banner.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), AZUL),
        ("TOPPADDING", (0,0), (-1,-1), 8),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
        ("LEFTPADDING", (0,0), (-1,-1), 12),
    ]))
    story.append(banner)
    story.append(Spacer(1, 8))

    # Info general
    if not info_sector.empty:
        story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
        story.append(Paragraph("Informaci\u00f3n General del Sector", estilo_seccion))
        row = info_sector.iloc[0]
        ig_data = [[Paragraph(str(k), estilo_label), Paragraph(str(v)[:100], estilo_normal)]
                   for k, v in row.items() if str(v) not in ("nan", "None", "")]
        if ig_data:
            t_ig = Table(ig_data, colWidths=["35%", "65%"])
            t_ig.setStyle(TableStyle([
                ("ROWBACKGROUNDS", (0,0), (-1,-1), [BLANCO, GRIS]),
                ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
                ("TOPPADDING", (0,0), (-1,-1), 4),
                ("BOTTOMPADDING", (0,0), (-1,-1), 4),
                ("LEFTPADDING", (0,0), (-1,-1), 6),
            ]))
            story.append(t_ig)

    # AOD
    if not aod_sector.empty:
        story.append(Spacer(1, 8))
        story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
        story.append(Paragraph("Ayuda Oficial al Desarrollo (AOD)", estilo_seccion))
        story.append(Paragraph("Fuente: C\u00edclope a corte de 26 de marzo de 2026", estilo_caption))
        aod_s2 = aod_sector.copy()
        aod_s2["VALOR APORTE (USD)"] = pd.to_numeric(aod_s2.get("VALOR APORTE (USD)", 0), errors="coerce").fillna(0)
        int_s = aod_s2["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in aod_s2.columns else 0
        coop_s = aod_s2["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in aod_s2.columns else 0
        usd_s = aod_s2["VALOR APORTE (USD)"].sum()
        metrics_s = [[
            Paragraph("INTERVENCIONES", estilo_label),
            Paragraph("COOPERANTES", estilo_label),
            Paragraph("TOTAL APORTE (USD)", estilo_label)
        ],[
            Paragraph(str(int_s), ParagraphStyle("v", fontName="Helvetica-Bold", fontSize=12, textColor=AZUL)),
            Paragraph(str(coop_s), ParagraphStyle("v", fontName="Helvetica-Bold", fontSize=12, textColor=AZUL)),
            Paragraph("USD " + f"{usd_s:,.0f}".replace(",", "."),
                      ParagraphStyle("v", fontName="Helvetica-Bold", fontSize=10, textColor=AZUL)),
        ]]
        t_m = Table(metrics_s, colWidths=["33%", "33%", "34%"])
        t_m.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), GRIS),
            ("LINEABOVE", (0,0), (-1,0), 3, AZUL),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 6),
            ("BOTTOMPADDING", (0,0), (-1,-1), 6),
            ("LEFTPADDING", (0,0), (-1,-1), 8),
        ]))
        story.append(t_m)

    # CSS
    if not css_sector.empty:
        story.append(Spacer(1, 8))
        story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
        story.append(Paragraph(f"Proyectos de Cooperaci\u00f3n Sur Sur ({len(css_sector)} proyectos)", estilo_seccion))
        COLS_CSS_PDF = ["C\u00f3digo", "VIA DE COOPERACION", "PAIS SOCIO", "REGION",
                        "NOMBRE DE LA INICIATIVA", "ESTADO"]
        cols_css = [c for c in COLS_CSS_PDF if c in css_sector.columns]
        css_data = [[Paragraph(c, estilo_label) for c in cols_css]]
        for _, r in css_sector[cols_css].head(20).iterrows():
            css_data.append([Paragraph(str(v)[:60], estilo_normal) for v in r.values])
        n = len(cols_css)
        t_css = Table(css_data, colWidths=[f"{100/n}%" for _ in range(n)])
        t_css.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 5),
            ("FONTSIZE", (0,0), (-1,-1), 6),
        ]))
        story.append(t_css)

    # ColCol
    if not colcol_sector.empty:
        story.append(Spacer(1, 8))
        story.append(HRFlowable(width="100%", thickness=2, color=AZUL, spaceAfter=4))
        story.append(Paragraph(f"Colombia Ense\u00f1a Colombia - ColCol ({len(colcol_sector)} intercambios)", estilo_seccion))
        COLS_CC_PDF = ["CODIGO", "NOMBRE DEL INTERCAMBIO", "A\u00d1O DE REALIZACI\u00d3N ",
                       "PRESUPUESTO ESTIMADO APC COLOMBIA"]
        cols_cc = [c for c in COLS_CC_PDF if c in colcol_sector.columns]
        cc_data = [[Paragraph(c, estilo_label) for c in cols_cc]]
        cc_show = colcol_sector[cols_cc].copy()
        if "PRESUPUESTO ESTIMADO APC COLOMBIA" in cc_show.columns:
            cc_show["PRESUPUESTO ESTIMADO APC COLOMBIA"] = (
                pd.to_numeric(cc_show["PRESUPUESTO ESTIMADO APC COLOMBIA"], errors="coerce")
                .apply(format_cop)
            )
        for _, r in cc_show.head(30).iterrows():
            cc_data.append([Paragraph(str(v)[:80], estilo_normal) for v in r.values])
        n = len(cols_cc)
        t_cc = Table(cc_data, colWidths=[f"{100/n}%" for _ in range(n)])
        t_cc.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), AZUL_CLARO),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [BLANCO, GRIS]),
            ("GRID", (0,0), (-1,-1), 0.5, GRIS_BORDE),
            ("TOPPADDING", (0,0), (-1,-1), 4),
            ("BOTTOMPADDING", (0,0), (-1,-1), 4),
            ("LEFTPADDING", (0,0), (-1,-1), 5),
            ("FONTSIZE", (0,0), (-1,-1), 6),
        ]))
        story.append(t_cc)

    # Footer
    story.append(Spacer(1, 14))
    story.append(HRFlowable(width="100%", thickness=0.5, color=GRIS_BORDE, spaceAfter=4))
    story.append(Paragraph(
        "Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia - APC-Colombia",
        estilo_caption))

    doc.build(story)
    output.seek(0)
    return output.getvalue()



# -------------------------
# APP
# -------------------------

# Load data
infogeneral, plan, ciclope, ciclope_ant, colcol, contrapartidas, proyectos, css = load_data()
info_s, aod_s, aod_s_ant, css_s, colcol_s = load_sectores()
geo = load_geo()

# \u2500\u2500 HEADER WITH LOGOS \u2500\u2500
col_title, col_logos = st.columns([3, 1])
with col_title:
    st.markdown(
        '<div style="padding: 0.6rem 0 0.2rem 0;">'
        '<div style="font-family:Montserrat,sans-serif;font-weight:800;font-size:1.25rem;color:#003087;line-height:1.3;">'
        'Caracterizaci\u00f3n territorial y sectorial para la gesti\u00f3n de la cooperaci\u00f3n internacional'
        '</div>'
        '<div style="font-size:0.82rem;color:#5A6A85;margin-top:4px;font-family:Source Sans 3,sans-serif;">'
        'Sistema Nacional de Cooperaci\u00f3n Internacional de Colombia'
        '</div></div>',
        unsafe_allow_html=True
    )
with col_logos:
    logo_col1, logo_col2 = st.columns(2)
    with logo_col1:
        try:
            st.image(LOGO_APC, use_container_width=True)
        except:
            st.markdown("APC")
    with logo_col2:
        try:
            st.image(LOGO_SNCIC, use_container_width=True)
        except:
            st.markdown("SNCIC")

st.markdown('<div style="height:4px;background:linear-gradient(90deg,#F5A623 33%,#003087 33% 66%,#C8102E 66%);margin-bottom:1.5rem;"></div>', unsafe_allow_html=True)

# \u2500\u2500 MAIN NAVIGATION \u2500\u2500
nav_options = ["\U0001f310 Panorama Nacional", "\U0001f5fa\ufe0f Ficha Territorial", "\U0001f3db\ufe0f Ficha Sectorial", "\U0001f4d6 Gu\u00eda de usuario"]
nav = st.radio("", nav_options, horizontal=True, label_visibility="collapsed",
               key="main_nav")

st.markdown("---")


DEPT_COL_INFO = "Departamento"
depts = sorted(infogeneral[DEPT_COL_INFO].dropna().unique().tolist())

# Pre-compute map data (interventions by dept for choropleth)
ciclope["DEPT_NORM"] = ciclope["DEPARTAMENTO"].map(norm_text)
ciclope_ant["DEPT_NORM"] = ciclope_ant["DEPARTAMENTO"].map(norm_text)
proyectos["DEPT_NORM"] = proyectos["DEPARTAMENTO"].map(norm_text)

# Build dept_interventions with norm_text keys to match GeoJSON
_di_raw = (
    ciclope[ciclope["DEPARTAMENTO"] != "\u00c1mbito Nacional"]
    .groupby("DEPARTAMENTO")["CODIGO INTERVENCION"].nunique()
)
dept_interventions = {norm_text(k): v for k, v in _di_raw.items()}


# =============================================================
# FICHA TERRITORIAL
# =============================================================
if nav == "\U0001f5fa\ufe0f Ficha Territorial":

    dept = st.selectbox("\U0001f5fa\ufe0f Selecciona un departamento", depts)

    infogeneral["DEPT_NORM"] = infogeneral[DEPT_COL_INFO].map(norm_text)
    dept_norm = norm_text(dept)
    info = infogeneral[infogeneral["DEPT_NORM"] == dept_norm].head(1)
    cic_dept = ciclope[ciclope["DEPT_NORM"] == dept_norm]
    cic_dept_ant = ciclope_ant[ciclope_ant["DEPT_NORM"] == dept_norm]
    proj_dept = proyectos[proyectos["DEPT_NORM"] == dept_norm]
    css_dept = css[css["ESPACIO VINCULADO"].map(norm_text) == dept_norm]

    mask_colcol = pd.Series(False, index=colcol.index)
    if "DEPARTAMENTOS PARTICIPANTES" in colcol.columns:
        mask_colcol = (
            colcol["DEPARTAMENTOS PARTICIPANTES"]
            .astype("string").map(norm_text)
            .str.contains(dept_norm, na=False)
        )
    colcol_dept = colcol[mask_colcol]

    if "Departamento" in contrapartidas.columns:
        contr_dept = contrapartidas[
            contrapartidas["Departamento"].astype("string").map(norm_text) == dept_norm
        ]
    else:
        contr_dept = contrapartidas.iloc[0:0]
    proj_dept_ant = ciclope_ant[ciclope_ant["DEPT_NORM"] == dept_norm]


    st.markdown(
        f'<div class="dept-title-banner">\U0001f4cd {dept}</div>',
        unsafe_allow_html=True
    )

    # ---- Informacion General ----
    st.markdown('<div class="section-header">Informaci\u00f3n General</div>', unsafe_allow_html=True)
    if info.empty:
        st.warning("No encontre el departamento en la tabla infogeneral.")
    else:
        c1, c2, c3 = st.columns(3)
        c1.metric("Capital", get_col(info.iloc[0], "Capital"))
        c2.metric("N\u00famero de Municipios",
                  get_col(info.iloc[0], "N\u00famero de Municipios", "Numero de Municipios", "Municipios"))
        pob_raw = get_col(info.iloc[0], "Poblaci\u00f3n", "Poblacion")
        try:
            pob_fmt = f"{int(float(pob_raw)):,}".replace(",", ".")
        except Exception:
            pob_fmt = str(pob_raw)
        c3.metric("Poblaci\u00f3n", pob_fmt)
        with st.expander("Ver registro completo del departamento"):
            df_det = info.T.reset_index()
            df_det.columns = ["Campo", "Valor"]
            df_det = df_det[~df_det["Campo"].astype(str).str.lower().str.strip().isin(
                ["porcentaje de avance", "dept_norm"])]
            st.dataframe(df_det, use_container_width=True, hide_index=True)

    # ---- AOD ----
    st.markdown('<div class="section-header">Ayuda Oficial al Desarrollo (AOD)</div>', unsafe_allow_html=True)

    m1, m2, m3, m4 = st.columns(4)
    int_26 = cic_dept["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in cic_dept.columns else 0
    int_25 = cic_dept_ant["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in cic_dept_ant.columns else 0
    cod_26 = set(cic_dept["CODIGO INTERVENCION"].dropna().unique())
    cod_25 = set(cic_dept_ant["CODIGO INTERVENCION"].dropna().unique())
    int_nuevas = len(cod_26 - cod_25)
    int_activas = len(cod_26 & cod_25)
    int_terminadas = len(cod_25 - cod_26)
    delta_int = int_26 - int_25
    delta_int_str = ("\u25b2 " if delta_int >= 0 else "\u25bc ") + str(abs(delta_int)) + " vs. 2026-1"
    delta_int_color = "#2E7D32" if delta_int >= 0 else "#C8102E"
    with m1:
        st.markdown(
            '<div class="metric-custom">'
            '<div class="metric-custom-label">Intervenciones (\u00fanicas)</div>'
            f'<div class="metric-custom-value">{int_26}</div>'
            f'<div class="metric-custom-delta" style="color:{delta_int_color};">{delta_int_str}</div>'
            f'<div class="metric-custom-sub">\u2665 {int_nuevas} nuevas &nbsp;|&nbsp; \u21ba {int_activas} contin\u00faan &nbsp;|&nbsp; \u2713 {int_terminadas} terminadas</div>'
            '</div>',
            unsafe_allow_html=True
        )
    coop_26 = cic_dept["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in cic_dept.columns else 0
    coop_ant = cic_dept_ant["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in cic_dept_ant.columns else 0
    delta_coop = coop_26 - coop_ant
    delta_coop_str = ("\u25b2 " if delta_coop >= 0 else "\u25bc ") + str(abs(delta_coop)) + " vs. 2026-1"
    delta_coop_color = "#2E7D32" if delta_coop >= 0 else "#C8102E"
    with m2:
        st.markdown(
            '<div class="metric-custom"><div class="metric-custom-label">Cooperantes</div>'
            f'<div class="metric-custom-value">{coop_26}</div>'
            f'<div class="metric-custom-delta" style="color:{delta_coop_color};">{delta_coop_str}</div></div>',
            unsafe_allow_html=True)
    municipios_count = (
        cic_dept["MUNICIPIO"].map(norm_text)
        .pipe(lambda s: s[~s.isin(["NO REPORTA", "SIN INFORMACION", "NO APLICA", ""])])
        .nunique() if "MUNICIPIO" in cic_dept.columns else 0
    )
    municipios_count_ant = (
        cic_dept_ant["MUNICIPIO"].map(norm_text)
        .pipe(lambda s: s[~s.isin(["NO REPORTA", "SIN INFORMACION", "NO APLICA", ""])])
        .nunique() if "MUNICIPIO" in cic_dept_ant.columns else 0
    )
    delta_mun = municipios_count - municipios_count_ant
    delta_mun_str = ("\u25b2 " if delta_mun >= 0 else "\u25bc ") + str(abs(delta_mun)) + " vs. 2026-1"
    delta_mun_color = "#2E7D32" if delta_mun >= 0 else "#C8102E"
    with m3:
        st.markdown(
            '<div class="metric-custom">'
            '<div class="metric-custom-label">Municipios o \u00e1reas no municipalizadas intervenidas</div>'
            f'<div class="metric-custom-value">{municipios_count}</div>'
            f'<div class="metric-custom-delta" style="color:{delta_mun_color};">{delta_mun_str}</div>'
            '</div>',
            unsafe_allow_html=True
        )
    total_usd = cic_dept["VALOR APORTE (USD)"].sum() if "VALOR APORTE (USD)" in cic_dept.columns else 0
    total_usd_ant = cic_dept_ant["VALOR APORTE (USD)"].sum() if "VALOR APORTE (USD)" in cic_dept_ant.columns else 0
    delta_usd = total_usd - total_usd_ant
    delta_usd_str = ("\u25b2 " if delta_usd >= 0 else "\u25bc ") + format_usd(abs(delta_usd)) + " vs. 2026-1"
    delta_usd_color = "#2E7D32" if delta_usd >= 0 else "#C8102E"
    with m4:
        st.markdown(
            '<div class="metric-custom">'
            '<div class="metric-custom-label">Total aporte estimado (USD)</div>'
            f'<div class="metric-custom-value" style="font-size:1.1rem;">{format_usd(total_usd)}</div>'
            f'<div class="metric-custom-delta" style="color:{delta_usd_color};">{delta_usd_str}</div>'
            '</div>',
            unsafe_allow_html=True
        )

    st.markdown("<br>", unsafe_allow_html=True)
    c5, c6 = st.columns(2)
    with c5:
        st.markdown("**Top 5 cooperantes por USD**")
        top_act = top_by_sum(cic_dept, "NOMBRE ACTOR", "VALOR APORTE (USD)", 5)
        if not top_act.empty:
            chart_act = (
                alt.Chart(top_act)
                .mark_bar(color="#003087", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("NOMBRE ACTOR:N", sort="-x", title=""),
                    x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                    tooltip=["NOMBRE ACTOR:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                ).properties(height=200)
            )
            st.altair_chart(chart_act, use_container_width=True)
            top_act_ant = top_by_sum(cic_dept_ant, "NOMBRE ACTOR", "VALOR APORTE (USD)", 5)
            top_act_disp = top_act.copy()
            top_act_disp.columns = ["NOMBRE ACTOR", "USD 2026-2"]
            top_act_disp["USD 2026-2"] = top_act_disp["USD 2026-2"].apply(format_usd)
            if not top_act_ant.empty:
                top_act_ant_disp = top_act_ant.copy()
                top_act_ant_disp.columns = ["NOMBRE ACTOR", "USD 2026-1"]
                top_act_ant_disp["USD 2026-1"] = top_act_ant_disp["USD 2026-1"].apply(format_usd)
                top_act_disp = top_act_disp.merge(top_act_ant_disp, on="NOMBRE ACTOR", how="left").fillna("-")
            st.dataframe(top_act_disp, use_container_width=True, hide_index=True)
        else:
            st.info("Sin datos suficientes para cooperantes.")
    with c6:
        st.markdown("**Top 5 ODS por USD**")
        top_ods = top_by_sum(cic_dept, "ODS", "VALOR APORTE (USD)", 5)
        if not top_ods.empty:
            chart_ods = (
                alt.Chart(top_ods)
                .mark_bar(color="#1565C0", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("ODS:N", sort="-x", title=""),
                    x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                    tooltip=["ODS:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                ).properties(height=200)
            )
            st.altair_chart(chart_ods, use_container_width=True)
            top_ods_ant = top_by_sum(cic_dept_ant, "ODS", "VALOR APORTE (USD)", 5)
            top_ods_disp = top_ods.copy()
            top_ods_disp.columns = ["ODS", "USD 2026-2"]
            top_ods_disp["ODS"] = top_ods_disp["ODS"].map(lambda x: ODS_NOMBRES.get(x, x))
            top_ods_disp["USD 2026-2"] = top_ods_disp["USD 2026-2"].apply(format_usd)
            if not top_ods_ant.empty:
                top_ods_ant_disp = top_ods_ant.copy()
                top_ods_ant_disp.columns = ["ODS", "USD 2026-1"]
                top_ods_ant_disp["ODS"] = top_ods_ant_disp["ODS"].map(lambda x: ODS_NOMBRES.get(x, x))
                top_ods_ant_disp["USD 2026-1"] = top_ods_ant_disp["USD 2026-1"].apply(format_usd)
                top_ods_disp = top_ods_disp.merge(top_ods_ant_disp, on="ODS", how="left").fillna("-")
            st.dataframe(top_ods_disp, use_container_width=True, hide_index=True)
        else:
            st.info("Sin datos suficientes para ODS.")

    # ---- Programas Internos APC ----
    st.markdown('<div class="section-header">Programas Internos APC-Colombia</div>', unsafe_allow_html=True)
    p1, p2 = st.columns(2)
    with p1:
        st.markdown("**ColCol - Colombia Ense\u00f1a Colombia**")
        colcol_unicos = colcol_dept["CODIGO"].nunique() if "CODIGO" in colcol_dept.columns else len(colcol_dept)
        st.metric("Intercambios \u00fanicos", colcol_unicos)
        COLS_CC = [
            "CODIGO", "ETAPA", "NOMBRE DEL INTERCAMBIO", "OBJETIVO DEL INTERCAMBIO",
            "BUENA PR\u00c1CTICA", "L\u00cdNEA TEM\u00c1TICA",
            "MUNICIPIO EN EL QUE SE DESARROLL\u00d3",
            "A\u00d1O DE REALIZACI\u00d3N ", "ENTIDAD SOCIA NACIONAL",
            "PRESUPUESTO ESTIMADO APC COLOMBIA"
        ]
        colcol_view = colcol_dept.copy()
        if "PRESUPUESTO ESTIMADO APC COLOMBIA" in colcol_view.columns:
            colcol_view["PRESUPUESTO ESTIMADO APC COLOMBIA"] = (
                pd.to_numeric(colcol_view["PRESUPUESTO ESTIMADO APC COLOMBIA"], errors="coerce")
                .apply(format_cop)
            )
        cols_cc_show = [c for c in COLS_CC if c in colcol_view.columns]
        st.dataframe(colcol_view[cols_cc_show], use_container_width=True, hide_index=True)
    with p2:
        st.markdown("**Contrapartidas**")
        st.metric("Registros encontrados", len(contr_dept))
        contr_view = contr_dept.copy()
        for col in contr_view.columns:
            if str(col).strip().strip("\'") in ["Monto por APC", "Monto total", "Monto total "]:
                contr_view[col] = pd.to_numeric(contr_view[col], errors="coerce").apply(format_cop)
        st.dataframe(contr_view.head(50), use_container_width=True, hide_index=True)

    # ---- CSS ----
    st.markdown('<div class="section-header">Proyectos de Cooperaci\u00f3n Sur Sur aprobados y vigentes</div>', unsafe_allow_html=True)
    st.caption("Datos actualizados a abril de 2026 \u00b7 APC Colombia, Direcci\u00f3n de Oferta")
    if css_dept.empty:
        st.info("No se encontraron proyectos de Cooperaci\u00f3n Sur Sur para este departamento.")
    else:
        proyectos_css_unicos = css_dept["C\u00f3digo"].nunique() if "C\u00f3digo" in css_dept.columns else len(css_dept)
        st.metric("Proyectos CSS \u00fanicos", proyectos_css_unicos)
        COLS_CSS = [
            "C\u00f3digo", "VIA DE COOPERACION", "MODALIDAD", "PAIS SOCIO", "SEGUNDO OFERENTE",
            "REGION", "NOMBRE DE LA INICIATIVA", "TIPO DE INICIATIVA", "FECHA DE APROBACION",
            "OBJETIVO GENERAL/DESCRIPCION DE LA INICIATIVA", "ESTADO",
            "ENTIDAD(ES) NACIONAL(ES)", "ENTIDAD(ES) EXTRANJERA(S)"
        ]
        cols_css_show = [c for c in COLS_CSS if c in css_dept.columns]
        css_disp = css_dept[cols_css_show].copy()
        st.dataframe(css_disp, use_container_width=True, hide_index=True)

    # ---- Proyectos AOD ----
    st.markdown('<div class="section-header">Proyectos AOD activos</div>', unsafe_allow_html=True)
    st.caption("Fuente: C\u00edclope a corte de 26 de marzo de 2026")
    df_aod_terr = proj_dept.drop(columns=["DEPT_NORM"], errors="ignore").copy()
    proy_unicos_terr = df_aod_terr["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in df_aod_terr.columns else len(df_aod_terr)
    proy_ant_t = proj_dept_ant["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in proj_dept_ant.columns else 0
    delta_pt = proy_unicos_terr - proy_ant_t
    delta_pt_str = ("\u25b2 " if delta_pt >= 0 else "\u25bc ") + str(abs(delta_pt)) + " vs. 2026-1"
    delta_pt_col = "#2E7D32" if delta_pt >= 0 else "#C8102E"
    st.markdown(
        '<div class="metric-custom" style="max-width:300px;">'
        '<div class="metric-custom-label">Proyectos AOD activos (\u00fanicos)</div>'
        f'<div class="metric-custom-value">{proy_unicos_terr}</div>'
        f'<div class="metric-custom-delta" style="color:{delta_pt_col};">{delta_pt_str}</div></div>',
        unsafe_allow_html=True)
    COLS_SHOW = ["CODIGO INTERVENCION", "NOMBRE INTERVENCION", "OBJETIVO GENERAL",
                 "FECHA INICIAL", "FECHA FINAL", "DEPARTAMENTO", "MUNICIPIO",
                 "NOMBRE ACTOR", "ENCI PRIMER NIVEL", "ODS", "SECTORES GOB"]
    cols_show = [c for c in COLS_SHOW if c in df_aod_terr.columns]
    st.dataframe(df_aod_terr[cols_show], use_container_width=True, hide_index=True)

    # ---- Descargas (al final, incluye todo) ----
    st.markdown("---")
    st.markdown("**Descargar ficha territorial completa**")
    excel_ficha = to_excel_ficha(info, cic_dept, colcol_dept, contr_dept, css_dept)
    col_pdf, col_xlsx = st.columns(2)
    with col_xlsx:
        st.download_button(
            label="\U0001f4e5 Descargar en Excel",
            data=excel_ficha,
            file_name=f"Ficha_Territorial_{dept}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_pdf:
        pdf_ficha = to_pdf_ficha(dept, info, cic_dept, colcol_dept, contr_dept, css_dept)
        st.download_button(
            label="\U0001f4e5 Descargar en PDF",
            data=pdf_ficha,
            file_name=f"Ficha_Territorial_{dept}.pdf",
            mime="application/pdf",
        )
    st.markdown('<div class="apc-footer">Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia \u00b7 APC-Colombia</div>', unsafe_allow_html=True)





# =============================================================
# FICHA SECTORIAL
# =============================================================
elif nav == "\U0001f3db\ufe0f Ficha Sectorial":

    sectores_list = info_s["Nombre del sector"].dropna().tolist()
    sector = st.selectbox("\U0001f3db\ufe0f Selecciona un sector", sectores_list)
    sector_norm = norm_text(sector)

    st.markdown(
        f'<div class="dept-title-banner">\U0001f3db\ufe0f {sector}</div>',
        unsafe_allow_html=True
    )

    # Info general del sector
    st.markdown('<div class="section-header">Informaci\u00f3n General del Sector</div>', unsafe_allow_html=True)
    info_sector = info_s[info_s["Nombre del sector"].map(norm_text) == sector_norm]
    if not info_sector.empty:
        row = info_sector.iloc[0]
        sg1, sg2 = st.columns(2)
        sg1.metric("Cabeza de sector", str(row.get("Cabeza de sector", "")))
        sg2.metric("Planes de trabajo vinculados al SNCIC",
                   "S\u00ed" if str(row.get("Cabezas del sector asociadas con planes de trabajo en el marco del SNCIC", "")).lower() in ["s\u00ed", "si", "yes"] else "No")
        planes = str(row.get("Planes de trabajo vinculados", ""))
        if planes and planes not in ("nan", "None", ""):
            st.markdown(f"**Planes de trabajo:** {planes}")

    # AOD del sector
    st.markdown('<div class="section-header">Ayuda Oficial al Desarrollo (AOD)</div>', unsafe_allow_html=True)
    st.caption("Fuente: C\u00edclope a corte de 26 de marzo de 2026")

    aod_sector = aod_s[aod_s["SECTORES GOB"].map(norm_text).str.contains(sector_norm, na=False)]

    if aod_sector.empty:
        st.info("No se encontraron intervenciones de AOD para este sector.")
    else:
        aod_sector["VALOR APORTE (USD)"] = pd.to_numeric(aod_sector["VALOR APORTE (USD)"], errors="coerce").fillna(0)
        aod_sector_ant = aod_s_ant[aod_s_ant["SECTORES GOB"].map(norm_text).str.contains(sector_norm, na=False)].copy()
        if "VALOR APORTE (USD)" in aod_sector_ant.columns:
            aod_sector_ant["VALOR APORTE (USD)"] = pd.to_numeric(aod_sector_ant["VALOR APORTE (USD)"], errors="coerce").fillna(0)
        int_s2 = aod_sector["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in aod_sector.columns else 0
        int_s1 = aod_sector_ant["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in aod_sector_ant.columns else 0
        d_int = int_s2 - int_s1
        d_int_str = ("\u25b2 " if d_int >= 0 else "\u25bc ") + str(abs(d_int)) + " vs. 2026-1"
        d_int_col = "#2E7D32" if d_int >= 0 else "#C8102E"
        coop_s2 = aod_sector["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in aod_sector.columns else 0
        coop_s1 = aod_sector_ant["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in aod_sector_ant.columns else 0
        d_coop = coop_s2 - coop_s1
        d_coop_str = ("\u25b2 " if d_coop >= 0 else "\u25bc ") + str(abs(d_coop)) + " vs. 2026-1"
        d_coop_col = "#2E7D32" if d_coop >= 0 else "#C8102E"
        total_s = aod_sector["VALOR APORTE (USD)"].sum()
        total_s1 = aod_sector_ant["VALOR APORTE (USD)"].sum() if "VALOR APORTE (USD)" in aod_sector_ant.columns else 0
        d_usd = total_s - total_s1
        d_usd_str = ("\u25b2 " if d_usd >= 0 else "\u25bc ") + format_usd(abs(d_usd)) + " vs. 2026-1"
        d_usd_col = "#2E7D32" if d_usd >= 0 else "#C8102E"
        sm1, sm2, sm3, sm4 = st.columns(4)
        with sm1:
            st.markdown('<div class="metric-custom"><div class="metric-custom-label">Intervenciones (\u00fanicas)</div>'
                f'<div class="metric-custom-value">{int_s2}</div>'
                f'<div class="metric-custom-delta" style="color:{d_int_col};">{d_int_str}</div></div>', unsafe_allow_html=True)
        with sm2:
            st.markdown('<div class="metric-custom"><div class="metric-custom-label">Cooperantes</div>'
                f'<div class="metric-custom-value">{coop_s2}</div>'
                f'<div class="metric-custom-delta" style="color:{d_coop_col};">{d_coop_str}</div></div>', unsafe_allow_html=True)
        sm3.metric("Departamentos", aod_sector[aod_sector["DEPARTAMENTO"] != "\u00c1mbito Nacional"]["DEPARTAMENTO"].nunique()
                   if "DEPARTAMENTO" in aod_sector.columns else 0)
        with sm4:
            st.markdown('<div class="metric-custom"><div class="metric-custom-label">Total aporte (USD)</div>'
                f'<div class="metric-custom-value" style="font-size:1rem;">{format_usd(total_s)}</div>'
                f'<div class="metric-custom-delta" style="color:{d_usd_col};">{d_usd_str}</div></div>', unsafe_allow_html=True)

        sc1, sc2 = st.columns(2)
        with sc1:
            st.markdown("**Top 5 cooperantes por USD**")
            top_coop_s = top_by_sum(aod_sector, "NOMBRE ACTOR", "VALOR APORTE (USD)", 5)
            if not top_coop_s.empty:
                chart_cs = (
                    alt.Chart(top_coop_s)
                    .mark_bar(color="#003087", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                    .encode(
                        y=alt.Y("NOMBRE ACTOR:N", sort="-x", title=""),
                        x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                        tooltip=["NOMBRE ACTOR:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                    ).properties(height=200)
                )
                st.altair_chart(chart_cs, use_container_width=True)
                top_coop_s_ant = top_by_sum(aod_sector_ant, "NOMBRE ACTOR", "VALOR APORTE (USD)", 5)
                top_coop_s_disp = top_coop_s.copy()
                top_coop_s_disp.columns = ["NOMBRE ACTOR", "USD 2026-2"]
                top_coop_s_disp["USD 2026-2"] = top_coop_s_disp["USD 2026-2"].apply(format_usd)
                if not top_coop_s_ant.empty:
                    tc1 = top_coop_s_ant.copy()
                    tc1.columns = ["NOMBRE ACTOR", "USD 2026-1"]
                    tc1["USD 2026-1"] = tc1["USD 2026-1"].apply(format_usd)
                    top_coop_s_disp = top_coop_s_disp.merge(tc1, on="NOMBRE ACTOR", how="left").fillna("-")
                st.dataframe(top_coop_s_disp, use_container_width=True, hide_index=True)

        with sc2:
            st.markdown("**Top 5 ODS por USD**")
            top_ods_s = top_by_sum(aod_sector, "ODS", "VALOR APORTE (USD)", 5)
            if not top_ods_s.empty:
                chart_os = (
                    alt.Chart(top_ods_s)
                    .mark_bar(color="#1565C0", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                    .encode(
                        y=alt.Y("ODS:N", sort="-x", title=""),
                        x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                        tooltip=["ODS:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                    ).properties(height=200)
                )
                st.altair_chart(chart_os, use_container_width=True)
                top_ods_s_ant = top_by_sum(aod_sector_ant, "ODS", "VALOR APORTE (USD)", 5)
                top_ods_s_disp = top_ods_s.copy()
                top_ods_s_disp.columns = ["ODS", "USD 2026-2"]
                top_ods_s_disp["ODS"] = top_ods_s_disp["ODS"].map(lambda x: ODS_NOMBRES.get(x, x))
                top_ods_s_disp["USD 2026-2"] = top_ods_s_disp["USD 2026-2"].apply(format_usd)
                if not top_ods_s_ant.empty:
                    to1 = top_ods_s_ant.copy()
                    to1.columns = ["ODS", "USD 2026-1"]
                    to1["ODS"] = to1["ODS"].map(lambda x: ODS_NOMBRES.get(x, x))
                    to1["USD 2026-1"] = to1["USD 2026-1"].apply(format_usd)
                    top_ods_s_disp = top_ods_s_disp.merge(to1, on="ODS", how="left").fillna("-")
                st.dataframe(top_ods_s_disp, use_container_width=True, hide_index=True)


    # CSS del sector
    st.markdown('<div class="section-header">Proyectos de Cooperaci\u00f3n Sur Sur</div>', unsafe_allow_html=True)
    st.caption("Datos actualizados a abril de 2026 \u00b7 APC Colombia, Direcci\u00f3n de Oferta")

    css_sector = css_s[css_s["ESPACIO VINCULADO"].map(norm_text).str.contains(sector_norm, na=False)]
    if css_sector.empty:
        st.info("No se encontraron proyectos CSS para este sector.")
    else:
        st.metric("Proyectos CSS \u00fanicos", css_sector["C\u00f3digo"].nunique() if "C\u00f3digo" in css_sector.columns else len(css_sector))
        COLS_CSS_S = [
            "C\u00f3digo", "VIA DE COOPERACION", "MODALIDAD", "PAIS SOCIO",
            "REGION", "NOMBRE DE LA INICIATIVA", "TIPO DE INICIATIVA",
            "FECHA DE APROBACION", "ESTADO", "ENTIDAD(ES) NACIONAL(ES)"
        ]
        cols_s = [c for c in COLS_CSS_S if c in css_sector.columns]
        st.dataframe(css_sector[cols_s], use_container_width=True, hide_index=True)

    # ColCol del sector
    st.markdown('<div class="section-header">Colombia Ense\u00f1a Colombia (ColCol)</div>', unsafe_allow_html=True)
    colcol_sector = colcol_s[colcol_s["SECTOR VINCULADO"].map(norm_text).str.contains(sector_norm, na=False)]
    if colcol_sector.empty:
        st.info("No se encontraron intercambios ColCol para este sector.")
    else:
        colcol_s_unicos = colcol_sector["CODIGO"].nunique() if "CODIGO" in colcol_sector.columns else len(colcol_sector)
        st.metric("Intercambios \u00fanicos", colcol_s_unicos)
        COLS_CC_S = [
            "CODIGO", "ETAPA", "NOMBRE DEL INTERCAMBIO", "OBJETIVO DEL INTERCAMBIO",
            "BUENA PR\u00c1CTICA", "L\u00cdNEA TEM\u00c1TICA",
            "MUNICIPIO EN EL QUE SE DESARROLL\u00d3",
            "A\u00d1O DE REALIZACI\u00d3N ", "ENTIDAD SOCIA NACIONAL",
            "PRESUPUESTO ESTIMADO APC COLOMBIA"
        ]
        cols_cc_s = [c for c in COLS_CC_S if c in colcol_sector.columns]
        colcol_s_disp = colcol_sector[cols_cc_s].copy()
        if "PRESUPUESTO ESTIMADO APC COLOMBIA" in colcol_s_disp.columns:
            colcol_s_disp["PRESUPUESTO ESTIMADO APC COLOMBIA"] = (
                pd.to_numeric(colcol_s_disp["PRESUPUESTO ESTIMADO APC COLOMBIA"], errors="coerce")
                .apply(format_cop)
            )
        st.dataframe(colcol_s_disp, use_container_width=True, hide_index=True)

    # ---- Proyectos AOD del sector ----
    st.markdown('<div class="section-header">Proyectos AOD del sector</div>', unsafe_allow_html=True)
    st.caption("Fuente: C\u00edclope a corte de 26 de marzo de 2026")
    aod_sector_proj = aod_s[aod_s["SECTORES GOB"].map(norm_text).str.contains(sector_norm, na=False)].copy()
    aod_sector_proj = aod_sector_proj.drop(columns=["DEPT_NORM"] if "DEPT_NORM" in aod_sector_proj.columns else [], errors="ignore")
    COLS_AOD_S = ["NOMBRE INTERVENCION", "OBJETIVO GENERAL", "FECHA INICIAL", "FECHA FINAL",
                  "DEPARTAMENTO", "MUNICIPIO", "NOMBRE ACTOR", "ENCI PRIMER NIVEL", "ODS", "SECTORES GOB"]
    cols_aod_s = [c for c in COLS_AOD_S if c in aod_sector_proj.columns]
    COLS_AOD_S = ["CODIGO INTERVENCION", "NOMBRE INTERVENCION", "OBJETIVO GENERAL",
                  "FECHA INICIAL", "FECHA FINAL", "DEPARTAMENTO", "MUNICIPIO",
                  "NOMBRE ACTOR", "ENCI PRIMER NIVEL", "ODS", "SECTORES GOB"]
    cols_aod_s = [c for c in COLS_AOD_S if c in aod_sector_proj.columns]
    if not aod_sector_proj.empty:
        proy_unicos_sect = aod_sector_proj["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in aod_sector_proj.columns else len(aod_sector_proj)
        aod_sp_ant = aod_s_ant[aod_s_ant["SECTORES GOB"].map(norm_text).str.contains(sector_norm, na=False)]
        proy_sect_ant = aod_sp_ant["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in aod_sp_ant.columns else 0
        d_ps = proy_unicos_sect - proy_sect_ant
        d_ps_str = ("\u25b2 " if d_ps >= 0 else "\u25bc ") + str(abs(d_ps)) + " vs. 2026-1"
        d_ps_col = "#2E7D32" if d_ps >= 0 else "#C8102E"
        st.markdown(
            '<div class="metric-custom" style="max-width:300px;">'
            '<div class="metric-custom-label">Proyectos AOD activos (\u00fanicos)</div>'
            f'<div class="metric-custom-value">{proy_unicos_sect}</div>'
            f'<div class="metric-custom-delta" style="color:{d_ps_col};">{d_ps_str}</div></div>',
            unsafe_allow_html=True)
        st.dataframe(aod_sector_proj[cols_aod_s], use_container_width=True, hide_index=True)
    else:
        st.info("Sin proyectos AOD para este sector.")

    # ---- Descargas Ficha Sectorial ----
    st.markdown("---")
    st.markdown("**Descargar ficha sectorial completa**")
    output_s = BytesIO()
    with pd.ExcelWriter(output_s, engine="openpyxl") as writer:
        if not info_sector.empty:
            info_sector.T.reset_index().rename(columns={"index": "Campo", 0: "Valor"}).to_excel(writer, sheet_name="Info General", index=False)
        if not aod_sector.empty:
            aod_sector.to_excel(writer, sheet_name="AOD", index=False)
        if not css_sector.empty:
            css_sector.to_excel(writer, sheet_name="CSS", index=False)
        if not colcol_sector.empty:
            colcol_sector.to_excel(writer, sheet_name="ColCol", index=False)
        if not aod_sector_proj.empty:
            aod_sector_proj.to_excel(writer, sheet_name="Proyectos AOD", index=False)
    output_s.seek(0)
    excel_sector = output_s.getvalue()
    col_s1, col_s2 = st.columns(2)
    with col_s1:
        st.download_button(
            label="\U0001f4e5 Descargar en Excel",
            data=excel_sector,
            file_name=f"Ficha_Sectorial_{sector}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    with col_s2:
        pdf_sector = to_pdf_sectorial(sector, info_sector, aod_sector, css_sector, colcol_sector)
        st.download_button(
            label="\U0001f4e5 Descargar en PDF",
            data=pdf_sector,
            file_name=f"Ficha_Sectorial_{sector}.pdf",
            mime="application/pdf",
        )

    st.markdown('<div class="apc-footer">Agencia Presidencial de Cooperaci\u00f3n Internacional de Colombia \u00b7 APC-Colombia</div>', unsafe_allow_html=True)



# =============================================================
# PANORAMA NACIONAL
# =============================================================
elif nav == "\U0001f310 Panorama Nacional":

    st.markdown(
        '<div class="dept-title-banner">\U0001f310 Panorama Nacional de la Cooperaci\u00f3n Internacional</div>',
        unsafe_allow_html=True
    )
    st.caption("Fuente: C\u00edclope a corte de 26 de marzo de 2026. Incluye \u00e1mbito nacional y territorial.")

    # Calcular datos nacionales
    cic_nacional = ciclope.copy()
    cic_nacional["VALOR APORTE (USD)"] = pd.to_numeric(cic_nacional["VALOR APORTE (USD)"], errors="coerce").fillna(0)

    cic_ant_nac = ciclope_ant.copy()
    cic_ant_nac["VALOR APORTE (USD)"] = pd.to_numeric(cic_ant_nac["VALOR APORTE (USD)"], errors="coerce").fillna(0)
    int_nac_26 = cic_nacional["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in cic_nacional.columns else 0
    int_nac_25 = cic_ant_nac["CODIGO INTERVENCION"].nunique() if "CODIGO INTERVENCION" in cic_ant_nac.columns else 0
    cod_nac_26 = set(cic_nacional["CODIGO INTERVENCION"].dropna().unique())
    cod_nac_25 = set(cic_ant_nac["CODIGO INTERVENCION"].dropna().unique())
    int_nac_nuevas = len(cod_nac_26 - cod_nac_25)
    int_nac_activas = len(cod_nac_26 & cod_nac_25)
    int_nac_terminadas = len(cod_nac_25 - cod_nac_26)
    delta_nac_int = int_nac_26 - int_nac_25
    delta_nac_int_str = ("\u25b2 " if delta_nac_int >= 0 else "\u25bc ") + str(abs(delta_nac_int)) + " vs. 2026-1"
    delta_nac_int_color = "#2E7D32" if delta_nac_int >= 0 else "#C8102E"
    total_nac = cic_nacional["VALOR APORTE (USD)"].sum()
    total_nac_fmt = "USD " + f"{total_nac/1_000_000:,.0f} M".replace(",", ".")
    total_nac_ant = cic_ant_nac["VALOR APORTE (USD)"].sum()
    delta_nac_usd = total_nac - total_nac_ant
    delta_nac_usd_str = ("\u25b2 " if delta_nac_usd >= 0 else "\u25bc ") + format_usd(abs(delta_nac_usd)) + " vs. 2026-1"
    delta_nac_usd_color = "#2E7D32" if delta_nac_usd >= 0 else "#C8102E"
    n1, n2, n3, n4 = st.columns(4)
    with n1:
        st.markdown(
            '<div class="metric-custom">'
            '<div class="metric-custom-label">Intervenciones (\u00fanicas)</div>'
            f'<div class="metric-custom-value">{int_nac_26}</div>'
            f'<div class="metric-custom-delta" style="color:{delta_nac_int_color};">{delta_nac_int_str}</div>'
            f'<div class="metric-custom-sub">\u2665 {int_nac_nuevas} nuevas &nbsp;|&nbsp; \u21ba {int_nac_activas} contin\u00faan &nbsp;|&nbsp; \u2713 {int_nac_terminadas} terminadas</div>'
            '</div>',
            unsafe_allow_html=True
        )
    coop_nac_26 = cic_nacional["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in cic_nacional.columns else 0
    coop_nac_ant = cic_ant_nac["NOMBRE ACTOR"].nunique() if "NOMBRE ACTOR" in cic_ant_nac.columns else 0
    d_coop_nac = coop_nac_26 - coop_nac_ant
    d_coop_nac_str = ("\u25b2 " if d_coop_nac >= 0 else "\u25bc ") + str(abs(d_coop_nac)) + " vs. 2026-1"
    d_coop_nac_col = "#2E7D32" if d_coop_nac >= 0 else "#C8102E"
    with n2:
        st.markdown(
            '<div class="metric-custom"><div class="metric-custom-label">Cooperantes</div>'
            f'<div class="metric-custom-value">{coop_nac_26}</div>'
            f'<div class="metric-custom-delta" style="color:{d_coop_nac_col};">{d_coop_nac_str}</div></div>',
            unsafe_allow_html=True)
    n3.metric("Departamentos con AOD",
              cic_nacional[cic_nacional["DEPARTAMENTO"] != "\u00c1mbito Nacional"]["DEPARTAMENTO"].nunique()
              if "DEPARTAMENTO" in cic_nacional.columns else 0)
    with n4:
        st.markdown(
            '<div class="metric-custom">'
            '<div class="metric-custom-label">Total aporte estimado (USD)</div>'
            f'<div class="metric-custom-value" style="font-size:1.1rem;">{total_nac_fmt}</div>'
            f'<div class="metric-custom-delta" style="color:{delta_nac_usd_color};">{delta_nac_usd_str}</div>'
            '</div>',
            unsafe_allow_html=True
        )

    st.markdown('<div class="section-header">Cooperantes</div>', unsafe_allow_html=True)
    c_n1, c_n2 = st.columns(2)

    with c_n1:
        st.markdown("**Top 10 cooperantes por recursos (USD)**")
        top_coop_usd = (
            cic_nacional.groupby("NOMBRE ACTOR")["VALOR APORTE (USD)"]
            .sum().sort_values(ascending=False).head(10).reset_index()
        )
        if not top_coop_usd.empty:
            chart_coop_usd = (
                alt.Chart(top_coop_usd)
                .mark_bar(color="#003087", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("NOMBRE ACTOR:N", sort="-x", title=""),
                    x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                    tooltip=["NOMBRE ACTOR:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                )
                .properties(height=280)
            )
            st.altair_chart(chart_coop_usd, use_container_width=True)
            top_coop_usd_disp = top_coop_usd.copy()
            top_coop_usd_disp["VALOR APORTE (USD)"] = top_coop_usd_disp["VALOR APORTE (USD)"].apply(format_usd)
            st.dataframe(top_coop_usd_disp, use_container_width=True, hide_index=True)

    with c_n2:
        st.markdown("**Top 10 cooperantes por n\u00famero de intervenciones**")
        top_coop_int = (
            cic_nacional.groupby("NOMBRE ACTOR")["CODIGO INTERVENCION"]
            .nunique().sort_values(ascending=False).head(10).reset_index()
        )
        top_coop_int.columns = ["NOMBRE ACTOR", "INTERVENCIONES"]
        if not top_coop_int.empty:
            chart_coop_int = (
                alt.Chart(top_coop_int)
                .mark_bar(color="#1565C0", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("NOMBRE ACTOR:N", sort="-x", title=""),
                    x=alt.X("INTERVENCIONES:Q", title="N\u00famero de intervenciones"),
                    tooltip=["NOMBRE ACTOR:N", "INTERVENCIONES:Q"]
                )
                .properties(height=280)
            )
            st.altair_chart(chart_coop_int, use_container_width=True)
            st.dataframe(top_coop_int, use_container_width=True, hide_index=True)

    st.markdown('<div class="section-header">ODS y Sectores</div>', unsafe_allow_html=True)
    c_n3, c_n4 = st.columns(2)

    with c_n3:
        st.markdown("**Top 10 ODS por recursos (USD)**")
        top_ods_nac = (
            cic_nacional.groupby("ODS")["VALOR APORTE (USD)"]
            .sum().sort_values(ascending=False).head(10).reset_index()
        )
        if not top_ods_nac.empty:
            chart_ods_nac = (
                alt.Chart(top_ods_nac)
                .mark_bar(color="#003087", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("ODS:N", sort="-x", title=""),
                    x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                    tooltip=["ODS:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                )
                .properties(height=280)
            )
            st.altair_chart(chart_ods_nac, use_container_width=True)
            top_ods_nac_disp = top_ods_nac.copy()
            top_ods_nac_disp["VALOR APORTE (USD)"] = top_ods_nac_disp["VALOR APORTE (USD)"].apply(format_usd)
            top_ods_nac_disp["ODS"] = top_ods_nac_disp["ODS"].map(lambda x: ODS_NOMBRES.get(x, x))
            st.dataframe(top_ods_nac_disp, use_container_width=True, hide_index=True)

    with c_n4:
        st.markdown("**Top 10 sectores por recursos (USD)**")
        top_sect_nac = (
            cic_nacional.groupby("SECTORES GOB")["VALOR APORTE (USD)"]
            .sum().sort_values(ascending=False).head(10).reset_index()
        )
        if not top_sect_nac.empty:
            chart_sect_nac = (
                alt.Chart(top_sect_nac)
                .mark_bar(color="#1565C0", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("SECTORES GOB:N", sort="-x", title=""),
                    x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                    tooltip=["SECTORES GOB:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                )
                .properties(height=280)
            )
            st.altair_chart(chart_sect_nac, use_container_width=True)
            top_sect_nac_disp = top_sect_nac.copy()
            top_sect_nac_disp["VALOR APORTE (USD)"] = top_sect_nac_disp["VALOR APORTE (USD)"].apply(format_usd)
            st.dataframe(top_sect_nac_disp, use_container_width=True, hide_index=True)

    st.markdown('<div class="section-header">Departamentos</div>', unsafe_allow_html=True)
    c_n5, c_n6 = st.columns(2)

    with c_n5:
        st.markdown("**Top 10 departamentos por recursos (USD)**")
        top_dept_usd = (
            cic_nacional[cic_nacional["DEPARTAMENTO"] != "\u00c1mbito Nacional"]
            .groupby("DEPARTAMENTO")["VALOR APORTE (USD)"]
            .sum().sort_values(ascending=False).head(10).reset_index()
        )
        if not top_dept_usd.empty:
            chart_dept_usd = (
                alt.Chart(top_dept_usd)
                .mark_bar(color="#003087", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("DEPARTAMENTO:N", sort="-x", title=""),
                    x=alt.X("VALOR APORTE (USD):Q", title="USD"),
                    tooltip=["DEPARTAMENTO:N", alt.Tooltip("VALOR APORTE (USD):Q", format=",.0f")]
                )
                .properties(height=280)
            )
            st.altair_chart(chart_dept_usd, use_container_width=True)
            top_dept_usd_disp = top_dept_usd.copy()
            top_dept_usd_disp["VALOR APORTE (USD)"] = top_dept_usd_disp["VALOR APORTE (USD)"].apply(format_usd)
            st.dataframe(top_dept_usd_disp, use_container_width=True, hide_index=True)

    with c_n6:
        st.markdown("**Top 10 departamentos por intervenciones**")
        top_dept_int = (
            cic_nacional[cic_nacional["DEPARTAMENTO"] != "\u00c1mbito Nacional"]
            .groupby("DEPARTAMENTO")["CODIGO INTERVENCION"]
            .nunique().sort_values(ascending=False).head(10).reset_index()
        )
        top_dept_int.columns = ["DEPARTAMENTO", "INTERVENCIONES"]
        if not top_dept_int.empty:
            chart_dept_int = (
                alt.Chart(top_dept_int)
                .mark_bar(color="#1565C0", cornerRadiusTopRight=4, cornerRadiusBottomRight=4)
                .encode(
                    y=alt.Y("DEPARTAMENTO:N", sort="-x", title=""),
                    x=alt.X("INTERVENCIONES:Q", title="N\u00famero de intervenciones"),
                    tooltip=["DEPARTAMENTO:N", "INTERVENCIONES:Q"]
                )
                .properties(height=280)
            )
            st.altair_chart(chart_dept_int, use_container_width=True)
            st.dataframe(top_dept_int, use_container_width=True, hide_index=True)

    st.markdown(
        '<div class="apc-footer">Agencia Presidencial de Cooperacion Internacional de Colombia - APC-Colombia</div>',
        unsafe_allow_html=True
    )



# =============================================================
# GUIA DE USUARIO
# =============================================================
elif nav == "\U0001f4d6 Gu\u00eda de usuario":

    st.markdown(
        '<div class="dept-title-banner">Gu\u00eda de usuario</div>',
        unsafe_allow_html=True
    )

    guia_html = (
        '<div class="guia-card">'
        '<div class="guia-intro">\u00bfQu\u00e9 es esta herramienta?</div>'
        '<p>Esta herramienta permite conocer c\u00f3mo se est\u00e1 moviendo la cooperaci\u00f3n '
        'internacional en Colombia, tanto desde una perspectiva territorial como sectorial. '
        'Integra informaci\u00f3n de la Ayuda Oficial al Desarrollo (AOD), la cooperaci\u00f3n Sur-Sur '
        'y los programas institucionales de APC-Colombia.</p>'
        '<p><strong>Panorama Nacional</strong> presenta los totales nacionales de intervenciones, '
        'cooperantes y recursos de cooperaci\u00f3n internacional, e identifica los principales '
        'cooperantes, ODS, sectores y departamentos con mayor financiaci\u00f3n. Incluye '
        'comparativos con el trimestre anterior.</p>'
        '<p><strong>Ficha Territorial</strong> permite explorar la cooperaci\u00f3n en cada uno '
        'de los 33 departamentos del pa\u00eds (incluida Bogot\u00e1, D.C.). Para cada territorio '
        'encontrar\u00e1: informaci\u00f3n general e institucional, indicadores de AOD con comparativos '
        'vs. el trimestre anterior, programas de la oferta de APC-Colombia (ColCol y Contrapartidas), '
        'proyectos de cooperaci\u00f3n Sur-Sur vigentes, y el listado detallado de proyectos AOD activos. '
        'Toda la informaci\u00f3n puede descargarse en Excel o PDF.</p>'
        '<p><strong>Ficha Sectorial</strong> permite explorar la cooperaci\u00f3n por sector de gobierno '
        '(26 sectores). Para cada sector encontrar\u00e1: informaci\u00f3n general, indicadores y '
        'gr\u00e1ficas de AOD, proyectos de cooperaci\u00f3n Sur-Sur, intercambios ColCol y el listado '
        'de proyectos AOD activos. La informaci\u00f3n puede descargarse en Excel o PDF.</p>'
        '<p>En algunos indicadores podr\u00e1 ver comparativos con el trimestre anterior. '
        'Las flechas \u25b2 (subi\u00f3) y \u25bc (baj\u00f3) indican la variaci\u00f3n respecto al per\u00edodo anterior. '
        'En la tarjeta de intervenciones: \u2665 nuevas &nbsp;|&nbsp; \u21ba contin\u00faan &nbsp;|&nbsp; \u2713 terminadas.</p>'
        '</div>'
    )
    st.markdown(guia_html, unsafe_allow_html=True)

    g1, g2, g3, g4 = st.columns(4)
    with g1:
        st.info("**\U0001f310 Panorama Nacional**\n\nVisi\u00f3n de la cooperaci\u00f3n en Colombia con comparativos.")
    with g2:
        st.info("**\U0001f5fa\ufe0f Ficha Territorial**\n\nCooperaci\u00f3n por departamento. Descarga en Excel y PDF.")
    with g3:
        st.info("**\U0001f3db\ufe0f Ficha Sectorial**\n\nCooperaci\u00f3n por sector de gobierno. Descarga en Excel y PDF.")
    with g4:
        st.info("**Fuentes**\n\nAOD: Sistema de Informaci\u00f3n C\u00edclope (10/07/2026). CSS: APC-Colombia, DOCI (04/2026). ColCol y Contrapartidas: APC-Colombia, DCI.")

    st.markdown(
        '<div class="apc-footer">Agencia Presidencial de Cooperacion Internacional de Colombia - APC-Colombia</div>',
        unsafe_allow_html=True
    )